# -*- coding: utf-8 -*-
# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1",   # CellRichText 는 3.1 부터다
#     "pillow>=10",      # 표지 부품 이미지. 없으면 그림 없이 발행된다
# ]
# ///
"""불량대책서 표준양식과 샘플을 한 정의에서 생성한다.

    uv run MAIN_SERVER/templates/build_report_template.py

의존성은 위 PEP 723 블록에 적혀 있어 uv 가 알아서 받는다. 이 저장소의 .venv 에는
openpyxl 이 없고, 문서 양식을 뽑자고 서버 환경에 넣을 이유도 없다.

두 파일이 따로 편집되면서 어긋나 있었다. 여기서 같이 만든다.
표준양식은 `{{token}}`을 그대로 두고, 샘플은 같은 자리에 값을 넣는다.

담당자 회신 칸은 **정의된 이름**(reply_*)으로 가리킨다. 좌표로 읽으면 담당자가
행을 하나만 삽입해도 파서가 엉뚱한 칸을 읽는다. 이름은 Excel이 따라 움직인다.

안내문은 셀 값이 아니라 **입력 메시지**로 붙인다. 셀에 안내문을 적어 두면
담당자가 지우지 않았을 때 그대로 결재에 올라가고, 파서는 그 문장을 원인으로 저장한다.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = Path(__file__).resolve().parent
REPO_ROOT = OUT_DIR.parent.parent
# 데이터시트 BOM 의 Image 열은 Unity 프로젝트의 Assets/ 기준 경로다.
UNITY_ASSETS = REPO_ROOT / "UnityDT" / "Assets"
TEMPLATE_PATH = OUT_DIR / "불량대책서_표준양식.xlsx"
SAMPLE_PATH = OUT_DIR / "불량대책서_표준양식_샘플.xlsx"
CLOSED_PATH = OUT_DIR / "불량대책서_샘플_종결본.xlsx"

# ---------------------------------------------------------------- 서식 상수
FACE = "맑은 고딕"
INK = "00262626"
MUTED = "00808080"
FAINT = "00A6A6A6"
BAND = "00404040"          # 섹션 머리 띠
SECTION = "00DCE3EF"       # 표 머리글
LABEL = "00F2F2F2"         # 라벨 칸
REPLY = "00FFFBF0"         # 담당자 입력 칸 — 크림색이 "여기가 내 칸"을 말한다
BANNER = "001F3864"        # 할 일 띠
ANALYSIS = "00EEF3FA"      # 자동 분석 본문
ANALYSIS_LABEL = "00DDE6F3"  # 자동 분석 라벨
ALERT = "00FCE9E9"         # 경고 배경
ALERT_INK = "009C2A2A"
NAVY = "001F3864"
PRIMARY_ROW = "00FFF6D9"

_thin = Side(style="thin", color=FAINT)
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
UNDER = Border(bottom=_thin)

# 타입 스케일 — 이 여섯 값 밖으로 나가지 않는다.
T_TITLE   = 18    # 문서 제목
T_FIGURE  = 12    # 현황 6개 수치
T_LABEL   = 9     # 라벨·값·섹션 띠
T_BODY    = 8.5   # 현상 문장
T_SMALL   = 8     # 회신·자동분석·표 데이터
T_MICRO   = 7.5   # 표 머리글·캡션·푸터

RATE_FMT = "0.000%"          # PPM 세계에서 0.00%는 자릿수가 모자란다
MULT_FMT = '0.0"배"'
PPM_FMT = '#,##0" PPM"'
QTY_FMT = "#,##0"
PP_FMT = '+0.00"%p";-0.00"%p"'

ALERT_STATUSES = "ISSUED,ASSIGNED,IN_PROGRESS,VERIFYING,CLOSED"
DEFECT_TYPES = "MISSING,POSITION_ERROR,ORIENTATION_ERROR,CRACK"
VERIFY_STATUSES = "PENDING,EFFECTIVE,INEFFECTIVE"

# ① 에서 담당자가 혼자 결정할 수 없는 조치들. 회신 시 ☐ 를 ☑ 로 바꾼다.
CONTAINMENT_SCOPE = ("협조 필요    ☐ 생산 중단    ☐ 출하 보류    "
                     "☐ 재고 선별    ☐ 고객 통보    ☐ 설비 정지")


def font(size: float = 9, bold: bool = False, color: str = INK, italic: bool = False) -> Font:
    return Font(name=FACE, size=size, bold=bold, color=color, italic=italic)


def fill(color):
    return PatternFill("solid", fgColor=color)


def align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def put(ws, ref, value, *, f=None, bg=None, al=None, border: Border | None = BOX,
        fmt=None, unlock=False):
    cell = ws[ref]
    cell.value = value
    cell.font = f or font()
    if bg:
        cell.fill = fill(bg)
    cell.alignment = al or align()
    if border is not None:
        cell.border = border
    if fmt:
        cell.number_format = fmt
    if unlock:
        cell.protection = Protection(locked=False)
    return cell


def label(ws, ref, text):
    return put(ws, ref, text, f=font(T_LABEL, True), bg=LABEL,
               al=align("left", "center", indent=1))


def band(ws, ref_range, text, note=""):
    """섹션 머리 띠. 검은 배경 + 흰 글씨."""
    ws.merge_cells(ref_range)
    first = ref_range.split(":")[0]
    cell = ws[first]
    cell.value = f"{text}        {note}" if note else text
    cell.font = font(T_LABEL, True, "00FFFFFF")
    cell.fill = fill(BAND)
    cell.alignment = align("left", "center", indent=1)


def table_head(ws, row, first_col, last_col, headers):
    for offset, text in enumerate(headers):
        cell = ws.cell(row=row, column=first_col + offset)
        cell.value = text
        cell.font = font(T_MICRO, True)
        cell.fill = fill(SECTION)
        cell.alignment = align("center", "center", wrap=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 24


def note_line(ws, ref_range, text, color=MUTED, size=7.5):
    ws.merge_cells(ref_range)
    cell = ws[ref_range.split(":")[0]]
    cell.value = text
    cell.font = font(size, False, color)
    cell.alignment = align("left", "center", wrap=True, indent=1)
    cell.border = None


def input_hint(ws, ref, title, prompt):
    """안내문을 셀 값이 아니라 입력 메시지로 붙인다."""
    dv = DataValidation(showInputMessage=True, showErrorMessage=False)
    dv.promptTitle = title
    dv.prompt = prompt
    ws.add_data_validation(dv)
    dv.add(ws[ref])


def list_rule(ws, ref, csv_values, title, prompt):
    dv = DataValidation(
        type="list",
        formula1=f'"{csv_values}"',
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    dv.promptTitle = title
    dv.prompt = prompt
    dv.errorTitle = "허용되지 않는 값"
    dv.error = f"다음 중 하나여야 한다: {csv_values}"
    ws.add_data_validation(dv)
    dv.add(ws[ref])


def page(ws, orientation="portrait", fit_height=0, gridlines=False):
    ws.sheet_view.showGridLines = gridlines
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.35
    ws.page_margins.top = ws.page_margins.bottom = 0.45


def place_part_image(ws, cell_ref, rel_path, px=42, dx=8, dy=10):
    """부품 렌더를 시트에 얹는다.

    담당자가 문서를 열었을 때 어떤 부품인지 글자를 읽기 전에 알아보게 하려는 것이다.
    경로는 데이터시트 BOM 의 Image 열(Assets/ 기준)에서 온다.
    Pillow 가 없거나 파일이 없으면 조용히 건너뛴다 — 그림 때문에 발행이 막히면 안 된다.
    """
    if not rel_path:
        return False
    src = UNITY_ASSETS / rel_path
    if not src.is_file():
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return False
    try:
        img = XLImage(str(src))
    except Exception:
        return False
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils import coordinate_to_tuple
    from openpyxl.utils.units import pixels_to_EMU

    row, col = coordinate_to_tuple(cell_ref)
    img.width = img.height = px
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=col - 1, colOff=pixels_to_EMU(dx),
                           row=row - 1, rowOff=pixels_to_EMU(dy)),
        ext=XDRPositiveSize2D(pixels_to_EMU(px), pixels_to_EMU(px)),
    )
    ws.add_image(img)
    return True


def lock_sheet(ws, allow_filter=True):
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    if allow_filter:
        ws.protection.autoFilter = False
        ws.protection.sort = False


# ---------------------------------------------------------------- 회신 칸 정의
# (정의된 이름, 셀, 안내 제목, 안내 본문) — DB 컬럼과 1:1
REPLY_CELLS = [
    (
        "reply_containment_scope", "대책서", "B21", "① 협조가 필요한 조치",
        "담당자 혼자 결정할 수 없는 항목이다. 해당하는 것의 ☐ 를 ☑ 로 바꾸고,\n"
        "아래 칸에 누구에게 언제 요청했는지 적는다.\n\n"
        "생산 중단 → 생산팀장 · 출하 보류 → 영업/물류 · 재고 선별 → 인력 투입 승인\n"
        "고객 통보 → 품질보증 책임자 · 설비 정지 → 설비팀\n"
        "→ containment_summary 첫 줄로 함께 저장된다",
    ),
    (
        "reply_containment", "대책서", "B22", "① 초동 조치 — 오늘 안에",
        "격리·선별·출하보류 범위와 완료 시각을 적는다.\n"
        "영향 받는 job_id / unit_id 범위를 함께 적는다.\n\n"
        "원인이 안 나와도 오늘 회신한다. 여기는 '막았는지'만 쓰는 칸이다.\n"
        "원인을 없애는 것은 ④ 이고, 이 칸은 확산을 멈추는 칸이다.\n"
        "→ alert_countermeasures.containment_summary",
    ),
    (
        "reply_root_cause_label", "대책서", "B25", "② 발생 원인 — 한 줄 요약 (40자)",
        "아래 본문을 한 줄로 줄인 것. 예: '배치 하강속도 상향 + 지지 핀 최원거리 슬롯'\n\n"
        "이 부품·유형이 다시 나면, 다음 대책서의 자동 분석 ③ 줄이 이 칸을 그대로 읽어\n"
        "'과거 원인'으로 싣는다. 본문은 길어서 그 자리에 들어가지 않는다.\n"
        "→ alert_countermeasures.root_cause_label",
    ),
    (
        "reply_root_cause", "대책서", "B26", "② 발생 원인 — 왜 만들어졌는가",
        "「판단자료」 A·B로 급증 시점과 레시피 변경 시점을 대조한다.\n"
        "재현 또는 실물 확인으로 검증한 근거를 적는다.\n"
        "가설이면 '가설'이라고 표시한다.\n"
        "→ alert_countermeasures.root_cause_summary",
    ),
    (
        "reply_escape_cause", "대책서", "B28", "③ 유출 원인 — 왜 검사에서 걸리지 않았는가",
        "「판단자료」 D의 검사 항목과 대조해 판정 기준의 한계를 적는다.\n"
        "검사를 안 한 것인지, 했는데 못 걸른 것인지 구분한다.\n"
        "→ alert_countermeasures.escape_cause_summary",
    ),
    (
        "reply_prevention", "대책서", "B31", "④ 재발방지 대책 — 원인을 제거하는 영구대책",
        "레시피·공정조건을 바꾸면 변경 후 recipe_version을 적는다.\n"
        "부품을 바꾸면 「대체품」 시트의 대체코드를 적는다.\n"
        "동일 부품을 쓰는 다른 슬롯으로의 수평전개 여부를 포함한다.\n\n"
        "레시피 변경은 변경관리 승인이, 부품 교체는 고객 승인이 필요할 수 있다.\n"
        "승인 절차가 걸리면 그 일정도 함께 적는다.\n"
        "→ alert_countermeasures.applied_recipe_version / applied_at",
    ),
    (
        "reply_verify_period", "검증", "B5", "⑤ 검증 기간",
        "대책을 적용하고 다시 집계한 구간. [시작, 끝) 으로 적는다.\n"
        "검증 물량이 기준 최소 건수를 넘어야 판정할 수 있다.",
    ),
    (
        "reply_verify_recipe", "검증", "D5", "⑤ 적용 레시피",
        "대책을 적용한 뒤 실행한 recipe_version.\n"
        "이 값이 있어야 적용 전후를 같은 제품 안에서 비교할 수 있다.\n"
        "→ alert_countermeasures.applied_recipe_version",
    ),
    (
        "reply_verify_inspected", "검증", "E5", "⑤ 검증 검사 수량",
        "부품 기준 검사 건수 (유닛 수 × 해당 부품 슬롯 수).\n"
        "→ alert_countermeasures.verified_inspected_quantity",
    ),
    (
        "reply_verify_defective", "검증", "F5", "⑤ 검증 불량 수량",
        "→ alert_countermeasures.verified_defective_quantity",
    ),
    (
        "reply_verify_status", "검증", "H5", "⑤ 판정",
        "기준 불량률 이하로 내려왔으면 EFFECTIVE.\n"
        "아니면 INEFFECTIVE로 두고 ②로 되돌아간다.\n"
        "→ alert_countermeasures.verification_status",
    ),
    (
        "reply_verify_note", "검증", "B6", "⑤ 판정 근거",
        "재집계 결과를 그렇게 읽은 이유.\n"
        "검증 수량이 기준 최소 건수에 못 미치면 그것도 적는다.",
    ),
    (
        "reply_closed_at", "검증", "F8", "⑥ 종결일",
        "→ alerts.closed_at · alert_status='CLOSED'",
    ),
    (
        "reply_closure_note", "검증", "B9", "⑥ 종결 의견",
        "잔여 위험과 종결 승인 의견.\n"
        "→ alert_countermeasures.closure_note / closed_by",
    ),
]


def _rich(*parts):
    """(텍스트, InlineFont) 조각을 한 셀에 섞어 넣는다."""
    return CellRichText(*[TextBlock(f, t) for t, f in parts])


def _inline(size=8, bold=False, color=INK, italic=False):
    return InlineFont(rFont=FACE, sz=size, b=bold, i=italic, color=color[-6:])


# ---------------------------------------------------------------- 시트: 대책서
def build_cover(wb, d):
    """담당자가 열었을 때 '내가 쓸 칸이 몇 개고 언제까지인지'가 먼저 보이게 짠다.

    수치를 위에 쌓아 두면 읽는 사람이 자기 할 일을 찾으려고 스무 줄을 훑어야 한다.
    그래서 순서를 이렇게 둔다 —
      머리(누가·무엇) → **할 일 띠** → 현황(가로 한 줄) → 시스템 분석 → 담당자 작성.

    불량률은 % 하나로 보여주지 않는다. 배치 불량은 PPM으로 세고, 슬롯이 25개인
    보드에서는 부품 하나의 불량률이 완제품에서 증폭된다. 그래서 한 줄에
    **불량률 · 기준 대비 배수 · PPM · 완제품 영향**을 함께 싣는다.
    """
    ws = wb.create_sheet("대책서")
    for col, w in {"A": 14, "B": 12, "C": 12, "D": 11,
                   "E": 14, "F": 12, "G": 11, "H": 11}.items():
        ws.column_dimensions[col].width = w
    heights = {
        1: 15, 2: 17, 3: 17, 4: 6,
        5: 19, 6: 26, 7: 19, 8: 6,
        9: 23, 10: 17, 11: 19, 12: 15, 13: 21, 14: 19,
        15: 17, 16: 19, 17: 19, 18: 19, 19: 19,
        20: 17,
        21: 16, 22: 18, 23: 18, 24: 18,
        25: 16, 26: 19, 27: 19,
        28: 18, 29: 18, 30: 18,
        31: 18, 32: 18, 33: 18,
        34: 12,
    }
    for row, h in heights.items():
        ws.row_dimensions[row].height = h

    # ── 제목 · 결재란 ────────────────────────────────────────────────
    # 부품 렌더에 전용 칸을 준다. 제목 위에 떠 있으면 로고처럼 보이고, 무엇을 가리키는지
    # 애매하다. 칸을 나눠 두면 "이 문서가 다루는 부품"이라는 자리값이 생긴다.
    ws.merge_cells("A1:A3")
    put(ws, "A1", None, bg=LABEL)
    place_part_image(ws, "A1", d.get("part_image"), px=46, dx=29, dy=10)

    ws.merge_cells("B1:D3")
    title = ws["B1"]
    title.value = "불  량  대  책  서"
    title.font = Font(name=FACE, size=T_TITLE, bold=True, color=INK)
    title.alignment = align("center", "center")
    title.border = BOX

    ws.merge_cells("E1:E3")
    put(ws, "E1", "결\n재", f=font(9, True), bg=LABEL,
        al=align("center", "center", wrap=True))
    for col, head in (("F", "작성"), ("G", "검토"), ("H", "승인")):
        put(ws, f"{col}1", head, f=font(8, True), bg=LABEL, al=align("center", "center"))
        ws.merge_cells(f"{col}2:{col}3")
        put(ws, f"{col}2", None, al=align("center", "center"))

    # ── 문서 머리 ───────────────────────────────────────────────────
    # 초동조치는 담당자 혼자 결정할 수 없는 항목을 포함한다. 협조 부서를 문서 머리에
    # 둬서, 회신 전에 누구를 끌어와야 하는지가 먼저 보이게 한다.
    for row, l_text, l_key, r_text, r_key in (
        (5, "문서번호", "alert_code", "발행일시", "issued_at"),
        (6, "대상 부품", "part_line", "불량 유형", "defect_type"),
        (7, "주관 담당", "assignee", "협조 부서", "support_teams"),
    ):
        label(ws, f"A{row}", l_text)
        ws.merge_cells(f"B{row}:D{row}")
        put(ws, f"B{row}", d[l_key],
            al=align("left", "center", wrap=(row == 6), indent=1))
        label(ws, f"E{row}", r_text)
        ws.merge_cells(f"F{row}:H{row}")
        put(ws, f"F{row}", d[r_key], al=align("left", "center", indent=1))

    list_rule(ws, "F6", DEFECT_TYPES, "불량 유형",
              "별지 3장은 유형과 무관하게 항상 첨부한다.\n"
              "어느 자료가 이번 건에 쓸모 있는지는 담당자가 판단한다.")

    # ── 할 일 띠 — 이 문서에서 담당자가 제일 먼저 읽어야 할 한 줄 ──────
    ws.merge_cells("A9:H9")
    todo = ws["A9"]
    if d["alert_status"] == "CLOSED":
        # 종결된 문서는 남은 할 일이 없다. 결과를 대신 싣는다.
        todo.value = _rich(
            (f"  {d['alert_status']}  ", _inline(7.5, True, "001F3864")),
            ("   ①~④ 작성 완료   ", _inline(8, False, "00E8ECF2")),
            ("⑤ 효과 검증 · ⑥ 종결은 「검증」 시트", _inline(8, True, "00E8ECF2")),
        )
    else:
        # 기한을 여기 한 줄로 모은다. 블록마다 흩어 두면 A열 라벨이 세 줄이 되고
        # 담당자는 일정을 여섯 군데서 주워 읽어야 한다.
        todo.value = _rich(
            (f"  {d['alert_status']}  ", _inline(7.5, True, "001F3864")),
            ("   회신 기한   ", _inline(7.5, False, "00A9B6C6")),
            ("①  ", _inline(9, True, "00FFD166")),
            (str(d["due_initial"]), _inline(9, True, "00FFD166")),
            ("     ②③  ", _inline(8, True, "00E8ECF2")),
            (str(d["due_cause"]), _inline(8, True, "00E8ECF2")),
            ("     ④  ", _inline(8, True, "00E8ECF2")),
            (str(d["due_action"]), _inline(8, True, "00E8ECF2")),
            ("     ⑤⑥ 「검증」 시트", _inline(7.5, False, "00A9B6C6")),
        )
    todo.fill = fill(BANNER)
    todo.alignment = align("center", "center")
    todo.border = BOX

    # ── 1. 불량 현황 ────────────────────────────────────────────────
    band(ws, "A10:H10", "1.  불량 현황")

    label(ws, "A11", "집계 기간")
    ws.merge_cells("B11:D11")
    put(ws, "B11", d["period_line"], al=align("left", "center", indent=1))
    label(ws, "E11", "레시피 버전")
    ws.merge_cells("F11:H11")
    put(ws, "F11", d["source_recipe_version"], al=align("left", "center", indent=1))

    # 다섯 값을 가로 한 줄에 편다. 기준·실제 불량률(%)은 뺐다 — PPM 과 배수가
    # 같은 값을 이미 두 번 말하고 있고, 이 문서는 PPM 으로 읽기로 했다.
    ws.merge_cells("C12:D12")
    ws.merge_cells("F12:H12")
    for ref, text in (("A12", "검사 수량"), ("B12", "불량 수량"),
                      ("C12", d["ppm_header"]), ("E12", "기준 대비"),
                      ("F12", "완제품 영향")):
        put(ws, ref, text, f=font(T_MICRO, True, MUTED), bg=LABEL,
            al=align("center", "center", wrap=True))

    put(ws, "A13", d["inspected_quantity"], f=font(T_FIGURE - 1.5),
        al=align("center", "center"), fmt=QTY_FMT)
    put(ws, "B13", d["defective_quantity"], f=font(T_FIGURE - 1.5),
        al=align("center", "center"), fmt=QTY_FMT)
    ws.merge_cells("C13:D13")
    put(ws, "C13", d["defect_ppm"], f=font(T_FIGURE, True, ALERT_INK), bg=ALERT,
        al=align("center", "center"), fmt=QTY_FMT)
    put(ws, "E13", d["vs_threshold_ratio"], f=font(T_FIGURE, True, ALERT_INK),
        bg=ALERT, al=align("center", "center"), fmt=MULT_FMT)
    ws.merge_cells("F13:H13")
    put(ws, "F13", d["unit_impact"], f=font(T_BODY, True), al=align("center", "center"))

    label(ws, "A14", "발생 집중")
    ws.merge_cells("B14:D14")
    put(ws, "B14", d["hotspot"], al=align("left", "center", indent=1))
    label(ws, "E14", "직전 기간 대비")
    ws.merge_cells("F14:H14")
    put(ws, "F14", d["trend_vs_prev"], f=font(T_BODY),
        al=align("left", "center", indent=1))

    # ── 2. 시스템이 먼저 본 것 ──────────────────────────────────────
    band(ws, "A15:H15", "2.  시스템이 먼저 본 것")

    ws.merge_cells("A16:A19")
    put(ws, "A16", "자동 분석", f=font(9, True, NAVY), bg=ANALYSIS_LABEL,
        al=align("center", "center", wrap=True))
    ws.merge_cells("B16:H19")
    put(ws, "B16", d["auto_analysis"], f=font(T_SMALL, False, NAVY), bg=ANALYSIS,
        al=align("left", "top", wrap=True, indent=1))

    # ── 3. 담당자 작성 ──────────────────────────────────────────────
    band(ws, "A20:H20", "3.  담당자 작성")

    # ① 은 네 줄이다. 첫 줄은 혼자 결정할 수 없는 조치의 체크 라인.
    ws.merge_cells("A21:A24")
    cell = ws["A21"]
    cell.value = _rich(
        ("①\n", _inline(11, True, INK)),
        ("초동 조치", _inline(8.5, True, INK)),
    )
    cell.font = font(9, True)
    cell.fill = fill(LABEL)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = BOX

    ws.merge_cells("B21:H21")
    put(ws, "B21", d.get("containment_scope", CONTAINMENT_SCOPE), bg=REPLY,
        f=font(8, True, MUTED), al=align("left", "center", indent=1), unlock=True)
    ws.merge_cells("B22:H24")
    put(ws, "B22", d.get("reply_22"), bg=REPLY, f=font(T_SMALL),
        al=align("left", "top", wrap=True, indent=1), unlock=True)

    # ② 만 회신 칸이 둘이다. 첫 줄은 다음 대책서의 자동 분석 ③ 줄이 읽어 갈 한 줄 요약.
    # 본문을 그 자리에 넣으면 문장이 문단째로 들어가므로, 짧게 따로 받는다.
    for row, num, name, split in (
        (25, "②", "발생 원인", True),
        (28, "③", "유출 원인", False),
        (31, "④", "재발방지\n대책", False),
    ):
        ws.merge_cells(f"A{row}:A{row + 2}")
        cell = ws[f"A{row}"]
        cell.value = _rich(
            (f"{num}\n", _inline(11, True, INK)),
            (f"{name}", _inline(8.5, True, INK)),
        )
        cell.font = font(9, True)
        cell.fill = fill(LABEL)
        cell.alignment = align("center", "center", wrap=True)
        cell.border = BOX
        if split:
            ws.merge_cells(f"B{row}:H{row}")
            put(ws, f"B{row}", d.get(f"reply_{row}"), bg=REPLY,
                f=font(T_SMALL, True), al=align("left", "center", indent=1),
                unlock=True)
            ws.merge_cells(f"B{row + 1}:H{row + 2}")
            put(ws, f"B{row + 1}", d.get(f"reply_{row + 1}"), bg=REPLY,
                f=font(T_SMALL), al=align("left", "top", wrap=True, indent=1),
                unlock=True)
        else:
            ws.merge_cells(f"B{row}:H{row + 2}")
            put(ws, f"B{row}", d.get(f"reply_{row}"), bg=REPLY, f=font(T_SMALL),
                al=align("left", "top", wrap=True, indent=1), unlock=True)

    ws.merge_cells("A34:H34")
    put(ws, "A34", d["footer"], f=font(T_MICRO, False, MUTED),
        al=align("center", "center"), border=None)

    for name, sheet, ref, title_text, prompt in REPLY_CELLS:
        if sheet == "대책서":
            input_hint(ws, ref, title_text, prompt)

    ws.print_area = "A1:H34"
    page(ws, "portrait", fit_height=1)
    ws.oddFooter.left.text = "&9&K808080" + str(d["alert_code"])
    ws.oddFooter.right.text = "&9&K808080&P / &N"
    lock_sheet(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------- 시트: 검증
def build_verify(wb, d):
    """⑤ 효과 검증과 ⑥ 종결.

    표지에서 뗐다. ①~④ 는 발행 후 1~2주 안에 쓰지만, ⑤ 는 대책을 적용하고
    **다시 생산이 돌아야** 채울 수 있어 2~4주 뒤고 ⑥ 은 그 다음이다.
    작성 시점이 다른 칸을 한 장에 두면 담당자는 늘 절반이 빈 문서를 본다.

    회신 칸은 좌표가 아니라 정의된 이름으로 읽으므로, 시트가 갈려도 파서는 그대로다.
    """
    ws = wb.create_sheet("검증")
    for col, w in {"A": 14, "B": 12, "C": 12, "D": 11,
                   "E": 14, "F": 12, "G": 11, "H": 11}.items():
        ws.column_dimensions[col].width = w
    for row, h in {1: 20, 2: 15, 3: 8, 4: 15, 5: 21, 6: 34,
                   7: 8, 8: 19, 9: 34, 10: 12}.items():
        ws.row_dimensions[row].height = h

    band(ws, "A1:H1", "5 · 6.  효과 검증과 종결", "[담당자 작성]")
    note_line(ws, "A2:H2", d["verify_header"], color="00595959")

    # ⑤ 효과 검증
    ws.merge_cells("A4:A6")
    cell = ws["A4"]
    cell.value = _rich(
        ("⑤\n", _inline(11, True, INK)),
        ("효과 검증\n", _inline(8.5, True, INK)),
        (str(d["due_verify"]), _inline(7.5, True, MUTED)),
    )
    cell.font = font(9, True)
    cell.fill = fill(LABEL)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = BOX

    ws.merge_cells("B4:C4")
    for ref, text in (("B4", "검증 기간"), ("D4", "적용 레시피"), ("E4", "검증 수량"),
                      ("F4", "불량 수량"), ("G4", "검증 불량률"), ("H4", "판정")):
        put(ws, ref, text, f=font(T_MICRO, True, MUTED), bg=LABEL,
            al=align("center", "center", wrap=True))

    ws.merge_cells("B5:C5")
    put(ws, "B5", d.get("verify_period"), bg=REPLY, al=align("center", "center"),
        unlock=True)
    put(ws, "D5", d.get("verify_recipe"), bg=REPLY, al=align("center", "center"),
        unlock=True)
    put(ws, "E5", d.get("verify_inspected"), bg=REPLY, al=align("center", "center"),
        fmt=QTY_FMT, unlock=True)
    put(ws, "F5", d.get("verify_defective"), bg=REPLY, al=align("center", "center"),
        fmt=QTY_FMT, unlock=True)
    put(ws, "G5", '=IFERROR(F5/E5,"")', al=align("center", "center"), fmt=RATE_FMT)
    put(ws, "H5", d.get("verify_status", "PENDING"), bg=REPLY,
        al=align("center", "center"), unlock=True)
    list_rule(ws, "H5", VERIFY_STATUSES, "⑤ 판정",
              "기준 불량률 이하면 EFFECTIVE · 아니면 INEFFECTIVE로 두고 ②로 되돌아간다.")

    ws.merge_cells("B6:H6")
    put(ws, "B6", d.get("verify_note"), bg=REPLY, f=font(T_SMALL),
        al=align("left", "top", wrap=True, indent=1), unlock=True)

    # ⑥ 종결
    ws.merge_cells("A8:D8")
    put(ws, "A8", "⑥  종결 판정", f=font(9, True), bg=LABEL,
        al=align("left", "center", indent=1))
    label(ws, "E8", "종결일")
    ws.merge_cells("F8:H8")
    put(ws, "F8", d.get("closed_at"), bg=REPLY, al=align("center", "center"),
        unlock=True)

    put(ws, "A9", "종결 의견", f=font(9, True), bg=LABEL,
        al=align("left", "center", indent=1))
    ws.merge_cells("B9:H9")
    put(ws, "B9", d.get("closure_note"), bg=REPLY, f=font(T_SMALL),
        al=align("left", "top", wrap=True, indent=1), unlock=True)

    for name, sheet, ref, title_text, prompt in REPLY_CELLS:
        if sheet == "검증":
            input_hint(ws, ref, title_text, prompt)

    ws.print_area = "A1:H9"
    page(ws)
    lock_sheet(ws, allow_filter=False)
    return ws


# ---------------------------------------------------------------- 시트: 근거
def build_evidence(wb, d):
    ws = wb.create_sheet("근거")
    for col, w in {"A": 11, "B": 12, "C": 15, "D": 20, "E": 40}.items():
        ws.column_dimensions[col].width = w

    band(ws, "A1:E1", "발행 시점 고정 근거", "[발행 시점 고정]")
    note_line(ws, "A2:E2",
              f"production.unit_defects 확정 기록 · 문서번호 {d['alert_code']} · 수정 금지")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 15

    ws.merge_cells("A3:E3")
    put(ws, "A3", "슬롯별 발생 분포        어느 슬롯에 몰려 있는지 · 공정 원인 판단의 1차 근거",
        f=font(9, True), bg=SECTION, al=align("left", "center", indent=1))

    table_head(ws, 4, 1, 4, ["슬롯", "불량 건수", "슬롯 불량률", "비고"])
    ws.merge_cells("D4:E4")
    row = 5
    for slot in d["slot_rows"]:
        put(ws, f"A{row}", slot["slot_code"], al=align("center"))
        put(ws, f"B{row}", slot["count"], al=align("center"), fmt=QTY_FMT)
        put(ws, f"C{row}", slot["rate"], al=align("center"), fmt=RATE_FMT)
        ws.merge_cells(f"D{row}:E{row}")
        put(ws, f"D{row}", slot.get("note"), f=font(8), al=align("left", "center", indent=1))
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    put(ws, f"A{row}", "불량 유닛 목록", f=font(9, True), bg=SECTION,
        al=align("left", "center", indent=1))
    row += 1
    table_head(ws, row, 1, 5,
               ["unit_id", "슬롯", "불량 유형", "검사 시각", "검사 이미지 경로"])
    head_row = row
    row += 1
    for unit in d["evidence_rows"]:
        put(ws, f"A{row}", unit["unit_id"], al=align("center"))
        put(ws, f"B{row}", unit["slot_code"], al=align("center"))
        put(ws, f"C{row}", unit["defect_type"], al=align("center"))
        put(ws, f"D{row}", unit["inspected_at"], al=align("center"))
        put(ws, f"E{row}", unit["image_path"], f=font(8), al=align("left", "center", indent=1))
        ws.row_dimensions[row].height = 16
        row += 1

    ws.freeze_panes = f"A{head_row + 1}"
    page(ws, "portrait")
    lock_sheet(ws)
    return ws


# ---------------------------------------------------------------- 시트: 판단자료
def build_analysis(wb, d):
    ws = wb.create_sheet("판단자료")
    for col, w in {"A": 20, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13,
                   "G": 30, "H": 30}.items():
        ws.column_dimensions[col].width = w

    band(ws, "A1:H1", "판단 자료")
    note_line(ws, "A2:H2",
              "담당자가 ②발생 원인 · ③유출 원인을 좁히는 데 쓰는 참고 자료 · 자동 조회 · 결재 대상 아님")
    ws.merge_cells("A3:H3")
    put(ws, "A3",
        "판단 순서    ①「근거」어느 슬롯에 몰렸나  →  ② A 언제부터 올랐나  →  "
        "③ B 그때 무엇이 바뀌었나  →  ④ C 전에도 있었나  →  ⑤ D 무엇을 확인하나",
        f=font(8, True, NAVY), bg=LABEL, al=align("left", "center", indent=1))
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[3].height = 20

    row = 5

    def block(title, note):
        nonlocal row
        ws.merge_cells(f"A{row}:H{row}")
        put(ws, f"A{row}", title, f=font(9, True), bg=SECTION,
            al=align("left", "center", indent=1))
        ws.row_dimensions[row].height = 17
        row += 1
        note_line(ws, f"A{row}:H{row}", note)
        ws.row_dimensions[row].height = 14
        row += 1

    block("A.  주차별 추세        [발행 시점 고정]",
          "언제부터 올라갔는지 · 급변 시점과 아래 B의 레시피 변경 시점을 대조한다. "
          "주차 경계와 대책서 집계 구간은 일치하지 않는다.")
    table_head(ws, row, 1, 7,
               ["주차", "검사 수량", "불량 수량", "불량률", "기준 대비", "레시피", "비고"])
    ws.merge_cells(f"G{row}:H{row}")
    row += 1
    for r in d["weekly_rows"]:
        put(ws, f"A{row}", r["week"], al=align("center"))
        put(ws, f"B{row}", r["inspected"], al=align("center"), fmt=QTY_FMT)
        put(ws, f"C{row}", r["defective"], al=align("center"), fmt=QTY_FMT)
        put(ws, f"D{row}", r["rate"], al=align("center"), fmt=RATE_FMT)
        put(ws, f"E{row}", r["vs_threshold"], al=align("center"), fmt=MULT_FMT)
        put(ws, f"F{row}", r["recipe"], al=align("center"))
        ws.merge_cells(f"G{row}:H{row}")
        put(ws, f"G{row}", r.get("note"), f=font(8), al=align("left", "center", indent=1))
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    block("B.  레시피 변경 전후 불량률        [발행 시점 고정]",
          "datastation_part_rates · 공정 조건이 원인인지 부품이 원인인지 가르는 1차 근거")
    table_head(ws, row, 1, 7,
               ["레시피", "적용 기간", "검사 수량", "불량 수량", "불량률", "기준 대비", "변경 내용"])
    ws.merge_cells(f"G{row}:H{row}")
    row += 1
    for r in d["recipe_rows"]:
        put(ws, f"A{row}", r["recipe"], al=align("center"))
        put(ws, f"B{row}", r["period"], al=align("center"))
        put(ws, f"C{row}", r["inspected"], al=align("center"), fmt=QTY_FMT)
        put(ws, f"D{row}", r["defective"], al=align("center"), fmt=QTY_FMT)
        put(ws, f"E{row}", r["rate"], al=align("center"), fmt=RATE_FMT)
        put(ws, f"F{row}", r["vs_threshold"], al=align("center"), fmt=MULT_FMT)
        ws.merge_cells(f"G{row}:H{row}")
        put(ws, f"G{row}", r.get("change_note"), f=font(8),
            al=align("left", "top", wrap=True, indent=1))
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    block("C.  동일 부품 · 동일 유형 과거 대책서        [발행 시점 고정]",
          "datastation_alerts · 재발이면 이전 대책이 왜 듣지 않았는지부터 확인한다")
    table_head(ws, row, 1, 7,
               ["문서번호", "집계 기간", "불량률", "상태", "적용 대책", "",
                "당시 근본 원인"])
    ws.merge_cells(f"E{row}:F{row}")
    ws.merge_cells(f"G{row}:H{row}")
    row += 1
    for r in d["history_rows"]:
        put(ws, f"A{row}", r["alert_code"], f=font(8), al=align("center"))
        put(ws, f"B{row}", r["period"], f=font(8), al=align("center"))
        put(ws, f"C{row}", r["defect_rate"], f=font(8), al=align("center"), fmt=RATE_FMT)
        put(ws, f"D{row}", r["status"], f=font(8), al=align("center"))
        ws.merge_cells(f"E{row}:F{row}")
        put(ws, f"E{row}", r["applied"], f=font(8), al=align("center", "center", wrap=True))
        ws.merge_cells(f"G{row}:H{row}")
        put(ws, f"G{row}", r["root_cause"], f=font(8),
            al=align("left", "top", wrap=True, indent=1))
        ws.row_dimensions[row].height = 32
        row += 1

    row += 1
    block("D.  품질 체크리스트        [조회 시점 참조 · 데이터시트 갱신 시 바뀜]",
          "데이터시트 Checklist 시트 · 범주 = "
          "part_groups.category_label · 이 범주에서 무엇을 확인해야 하는지")
    table_head(ws, row, 1, 5, ["구분", "확인 항목", "", "", "이상 시 조치"])
    ws.merge_cells(f"B{row}:D{row}")
    ws.merge_cells(f"E{row}:H{row}")
    row += 1
    for key, title in [("incoming_inspection", "입고 검사"),
                       ("assembly_control", "조립·보관"),
                       ("reliability_test", "기능·신뢰성")]:
        put(ws, f"A{row}", title, f=font(8, True), bg=LABEL,
            al=align("left", "center", indent=1))
        ws.merge_cells(f"B{row}:D{row}")
        put(ws, f"B{row}", d["checklist"][key], f=font(8),
            al=align("left", "top", wrap=True, indent=1))
        if title == "입고 검사":
            ws.merge_cells(f"E{row}:H{row + 2}")
            put(ws, f"E{row}", d["checklist"]["action_on_anomaly"], f=font(8),
                al=align("left", "top", wrap=True, indent=1))
        ws.row_dimensions[row].height = 40
        row += 1

    page(ws, "landscape")
    lock_sheet(ws)
    return ws


# ---------------------------------------------------------------- 시트: 대체품
def build_alternates(wb, d):
    ws = wb.create_sheet("대체품")
    for col, w in {"A": 9, "B": 9, "C": 22, "D": 26, "E": 14, "F": 22,
                   "G": 29, "H": 13}.items():
        ws.column_dimensions[col].width = w

    band(ws, "A1:H1", "대체품 후보", "[조회 시점 참조]")
    note_line(ws, "A2:H2",
              "불량 유형과 무관하게 항상 첨부한다 · 부품 교체가 이번 건의 대책인지는 "
              "② 발생 원인을 확정한 담당자가 판단한다")
    ws.merge_cells("A3:H3")
    put(ws, "A3",
        "동일 정격만으로 drop-in 판단 금지 · 「필수 재검증」을 모두 통과하고 "
        "변경 승인이 완료된 뒤에만 적용한다.",
        f=font(8, True, ALERT_INK), bg=ALERT, al=align("left", "center", wrap=True, indent=1))
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[3].height = 22

    note_line(ws, "A4:H4", d["alternates_header"], color="00595959")
    ws.row_dimensions[4].height = 15

    table_head(ws, 6, 1, 8, [
        "역할", "대체코드", "제조사 / P/N", "핵심 사양·용도", "상호호환 상태",
        "필수 재검증", d["relevance_header"], "Source · 확인일",
    ])
    ws.row_dimensions[6].height = 26

    row = 7
    for cand in d["candidate_rows"]:
        is_primary = cand["role"] == "주품목"
        bg = PRIMARY_ROW if is_primary else None
        vals = [
            cand["role"], cand["alternate_code"], cand["maker_pn"], cand["key_spec"],
            cand["compatibility"], cand["revalidation"], cand["relevance"], cand["source"],
        ]
        for offset, value in enumerate(vals):
            cell = ws.cell(row=row, column=1 + offset)
            cell.value = value
            unrated = str(value).startswith("미평가")
            cell.font = font(8, is_primary, MUTED if unrated else INK, italic=unrated)
            cell.alignment = align("left", "top", wrap=True, indent=1)
            cell.border = BOX
            if bg:
                cell.fill = fill(bg)
        ws.row_dimensions[row].height = 40
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    put(ws, f"A{row}",
        "단가·MOQ·재고·리드타임·공급사는 이 문서에 싣지 않는다. 조회 시점에 이미 상하는 "
        "값이고 품질 담당자의 결정 변수가 아니다 → GET /parts/{part_id} 또는 구매 담당",
        f=font(7.5, False, MUTED), al=align("left", "center", wrap=True, indent=1), border=None)

    ws.freeze_panes = "C7"
    page(ws, "landscape", fit_height=1)
    lock_sheet(ws)
    return ws


# ---------------------------------------------------------------- 데이터
TOKEN_DATA = {
    "alert_code": "{{alert_code}}",
    "issued_at": "{{issued_at}}",
    "assignee": "{{assignee}}",
    "support_teams": "{{support_teams}}",
    "alert_status": "{{alert_status}}",
    "due_initial": "{{due_initial}}",
    "due_cause": "{{due_cause}}",
    "due_action": "{{due_action}}",
    "due_verify": "{{due_verify}}",
    "part_image": None,   # 표준양식에는 넣지 않는다 — 발행 시점에 부품이 정해진다
    "part_line": "{{part_id}} · {{part_name}} ({{group_id}} / {{category_label}})",
    "defect_type": "{{defect_type}}",
    "period_line": "{{period_start}} ~ {{period_end}} ({{window_days}}일 / {{evaluation_mode}})",
    "source_recipe_version": "{{source_recipe_version}}",
    "inspected_quantity": "{{inspected_quantity}}",
    "defective_quantity": "{{defective_quantity}}",
    "threshold_rate": "{{threshold_rate}}",
    "defect_rate": "{{defect_rate}}",
    "ppm_header": "실제 PPM (기준 {{threshold_ppm}})",
    "defect_ppm": "{{defect_ppm}}",
    "vs_threshold_ratio": "{{vs_threshold_ratio}}",
    "unit_impact": "{{unit_impact}}",
    "hotspot": "{{hotspot}}",
    "trend_vs_prev": "{{trend_vs_prev}}",
    "auto_analysis": "{{auto_analysis}}",
    "verify_header": "{{alert_code}}   ·   {{part_id}} / {{defect_type}}   ·   "
                     "발행 {{issued_at}}   ·   ④ 적용 후 다시 집계해서 채운다",
    "footer": "자동 발행   ·   회신 칸은 클릭하면 작성 요령이 뜬다",
    "slot_rows": [{"slot_code": "{{slot_code}}", "count": "{{count}}",
                   "rate": "{{slot_rate}}", "note": "{{slot_note}}"}],
    "evidence_rows": [{"unit_id": "{{unit_id}}", "slot_code": "{{slot_code}}",
                       "defect_type": "{{defect_type}}", "inspected_at": "{{inspected_at}}",
                       "image_path": "{{inspection_image_path}}"}],
    "weekly_rows": [{"week": "{{week}}", "inspected": "{{inspected}}",
                     "defective": "{{defective}}", "rate": "{{rate}}",
                     "vs_threshold": "{{vs_threshold}}", "recipe": "{{recipe_version}}",
                     "note": "{{note}}"}],
    "recipe_rows": [{"recipe": "{{recipe_version}}", "period": "{{applied_period}}",
                     "inspected": "{{inspected}}", "defective": "{{defective}}",
                     "rate": "{{rate}}", "vs_threshold": "{{vs_threshold}}",
                     "change_note": "{{change_note}}"}],
    "history_rows": [{"alert_code": "{{alert_code}}", "period": "{{period}}",
                      "defect_rate": "{{defect_rate}}", "status": "{{alert_status}}",
                      "applied": "{{applied_recipe_version}} / {{applied_at}}",
                      "root_cause": "{{root_cause_summary}}"}],
    "checklist": {
        "incoming_inspection": "{{incoming_inspection}}",
        "assembly_control": "{{assembly_control}}",
        "reliability_test": "{{reliability_test}}",
        "action_on_anomaly": "{{action_on_anomaly}}",
    },
    "alternates_header": "조회 시각 {{queried_at}}   ·   데이터시트 {{source_file}} "
                         "(기준일 {{source_dated_on}})   ·   Group ID {{group_id}} "
                         "({{category_label}})   ·   후보 {{candidate_count}}종   "
                         "·   행은 후보 수만큼 반복",
    "relevance_header": "{{defect_type}} 관련 소견",
    "candidate_rows": [{
        "role": "{{candidate_role}}", "alternate_code": "{{alternate_code}}",
        "maker_pn": "{{manufacturer}} / {{manufacturer_part_number}}",
        "key_spec": "{{key_spec}}", "compatibility": "{{compatibility_status}}",
        "revalidation": "{{revalidation_items}}", "relevance": "{{defect_relevance}}",
        "source": "{{source_id}} · {{checked_at}}",
    }],
}

# 샘플 수치는 004_mock_seed.sql 을 따른다.
# CAP 슬롯은 CAP-01~05, 유닛당 5개다. 검사 수량 = 완성품 2,500대 × 5 = 12,500.
#
# 기준은 400 PPM(0.04%)으로 잡았다. 슬롯이 25개인 보드에서 부품 단위 400 PPM 이면
# 완제품 수율은 (1-0.0004)^25 = 99.0% 다. 이전 샘플의 0.30% 기준은 완제품 기준으로
# 7.2% 불량이라 대책서를 쓸 상황이 아니라 라인을 세울 상황이었다.
BASE_UNITS = 2500
CAP_SLOTS = 5
INSPECTED = BASE_UNITS * CAP_SLOTS      # 12,500
DEFECTIVE = 16                          # 1,280 PPM
THRESHOLD = 0.0004                      # 400 PPM

SAMPLE_DATA = {
    "alert_code": "QA-CAP-CRACK-20260819-001",
    "issued_at": "2026-08-19 09:00:00",
    "assignee": "품질보증팀 / 배정 대기",
    "support_teams": "설비팀 · 생산팀 · 구매",
    "alert_status": "ISSUED",
    "due_initial": "오늘 18:00",
    "due_cause": "08-22 (3일)",
    "due_action": "09-05 (2주)",
    "due_verify": "적용 후 2~4주",
    "part_image": "UI/Icons/item-cap.png",
    "part_line": "CAP  (C-001 · MLCC)",
    "defect_type": "CRACK",
    "period_line": "2026-08-12 ~ 08-19  (7일 / ROLLING)",
    "source_recipe_version": "mock-v3",
    "inspected_quantity": INSPECTED,
    "defective_quantity": DEFECTIVE,
    "threshold_rate": THRESHOLD,
    "defect_rate": DEFECTIVE / INSPECTED,
    "ppm_header": f"실제 PPM (기준 {THRESHOLD * 1e6:,.0f})",
    "defect_ppm": DEFECTIVE / INSPECTED * 1e6,
    "vs_threshold_ratio": DEFECTIVE / INSPECTED / THRESHOLD,
    "unit_impact": "2,500대 중 15대 (0.60%)",
    "hotspot": "CAP-02 8건 · CAP-04 5건",
    "trend_vs_prev": "200 → 1,280 PPM  (6.4배 · 08-18 초과)",
    "auto_analysis": (
        "① 발생이 CAP-02·CAP-04에 81.3% 집중(판정 기준 60%) → 부품 Lot보다 공정·설비 원인을 "
        "먼저 의심한다.「근거」\n"
        "② 최초 초과(08-18)가 mock-v3 적용(08-16)과 3일 이내 → 변경 내용 "
        "'그리퍼 파지력 상향 · 배치 하강속도 상향'이 1순위 확인 대상.「판단자료」B\n"
        "③ 동일 부품·유형 CLOSED 대책서 1건(2026-04-12 · 배치 하강속도 과다) → 재발. "
        "당시 대책이 왜 듣지 않았는지 먼저 확인한다.「판단자료」C\n"
        "④ C-001 대체 후보 2종 · CRACK 관련 소견 2종.「대체품」"
    ),
    "verify_header": "QA-CAP-CRACK-20260819-001   ·   CAP / CRACK   ·   "
                     "발행 2026-08-19   ·   ④ 적용 후 다시 집계해서 채운다",
    "footer": "자동 발행   ·   샘플 (담당자 작성 영역 미작성)   ·   결재 대상은 이 시트 1장",
    "slot_rows": [
        {"slot_code": "CAP-02", "count": 8, "rate": 8 / BASE_UNITS,
         "note": "상위 2슬롯 81.3% · 공정/설비 우선 의심"},
        {"slot_code": "CAP-04", "count": 5, "rate": 5 / BASE_UNITS, "note": None},
        {"slot_code": "CAP-01", "count": 2, "rate": 2 / BASE_UNITS, "note": None},
        {"slot_code": "CAP-05", "count": 1, "rate": 1 / BASE_UNITS, "note": None},
    ],
    "evidence_rows": [
        {"unit_id": 10412, "slot_code": "CAP-02", "defect_type": "CRACK",
         "inspected_at": "2026-08-18 16:30:12",
         "image_path": "/var/lib/inspect/2026-08-18/u10412.png"},
        {"unit_id": 10418, "slot_code": "CAP-04", "defect_type": "CRACK",
         "inspected_at": "2026-08-18 16:41:55",
         "image_path": "/var/lib/inspect/2026-08-18/u10418.png"},
        {"unit_id": 10425, "slot_code": "CAP-02", "defect_type": "CRACK",
         "inspected_at": "2026-08-18 17:02:30",
         "image_path": "/var/lib/inspect/2026-08-18/u10425.png"},
        {"unit_id": 10431, "slot_code": "CAP-02", "defect_type": "CRACK",
         "inspected_at": "2026-08-18 17:19:08",
         "image_path": "/var/lib/inspect/2026-08-18/u10431.png"},
        {"unit_id": 10447, "slot_code": "CAP-04", "defect_type": "CRACK",
         "inspected_at": "2026-08-18 18:03:41",
         "image_path": "/var/lib/inspect/2026-08-18/u10447.png"},
        {"unit_id": "…", "slot_code": "…", "defect_type": "…", "inspected_at": "…",
         "image_path": "총 16행 · 발행 시점 스냅샷"},
    ],
    "weekly_rows": [
        {"week": "2026-W29", "inspected": 12400, "defective": 2, "rate": 2 / 12400,
         "vs_threshold": (2 / 12400) / THRESHOLD, "recipe": "mock-v1", "note": None},
        {"week": "2026-W30", "inspected": 12550, "defective": 3, "rate": 3 / 12550,
         "vs_threshold": (3 / 12550) / THRESHOLD, "recipe": "mock-v1", "note": None},
        {"week": "2026-W31", "inspected": 12480, "defective": 2, "rate": 2 / 12480,
         "vs_threshold": (2 / 12480) / THRESHOLD, "recipe": "mock-v2",
         "note": "08-05 mock-v2 적용"},
        {"week": "2026-W32", "inspected": 12575, "defective": 3, "rate": 3 / 12575,
         "vs_threshold": (3 / 12575) / THRESHOLD, "recipe": "mock-v2", "note": None},
        {"week": "2026-W33", "inspected": 12650, "defective": 8, "rate": 8 / 12650,
         "vs_threshold": (8 / 12650) / THRESHOLD, "recipe": "mock-v3",
         "note": "08-16 mock-v3 적용 · 급증 시점"},
        {"week": "2026-W34", "inspected": 5420, "defective": 8, "rate": 8 / 5420,
         "vs_threshold": (8 / 5420) / THRESHOLD, "recipe": "mock-v3",
         "note": "진행 중 · 08-19 09:00 기준"},
    ],
    "recipe_rows": [
        {"recipe": "mock-v1", "period": "~ 2026-08-05", "inspected": 124000,
         "defective": 20, "rate": 20 / 124000,
         "vs_threshold": (20 / 124000) / THRESHOLD, "change_note": "기준 레시피"},
        {"recipe": "mock-v2", "period": "08-05 ~ 08-16", "inspected": 99500,
         "defective": 24, "rate": 24 / 99500,
         "vs_threshold": (24 / 99500) / THRESHOLD, "change_note": "픽업 속도 조정"},
        {"recipe": "mock-v3", "period": "08-16 ~ 현재", "inspected": 30600,
         "defective": 33, "rate": 33 / 30600,
         "vs_threshold": (33 / 30600) / THRESHOLD,
         "change_note": "그리퍼 파지력 상향 · 배치 하강속도 상향 — 변경 직후 CRACK 급증"},
    ],
    "history_rows": [
        {"alert_code": "QA-CAP-CRACK-20260412-001", "period": "2026-04-05 ~ 04-12",
         "defect_rate": 0.0009, "status": "CLOSED", "applied": "mock-v1 / 04-15",
         "root_cause": "배치 하강속도 과다로 착지 충격. 속도 하향 후 160 PPM으로 회복."},
    ],
    "checklist": {
        "incoming_inspection": "외관/단자, 0603·100V·X7R 확인, LCR·절연저항 샘플",
        "assembly_control": "land/reflow 준수; panel depanel·ICT probe·connector 체결 시 "
                            "board support로 보드 굽힘 억제",
        "reliability_test": "AOI, LCR/IR; crack 의심 Lot 단면 또는 비파괴 검사",
        "action_on_anomaly": "크랙/IR 저하는 Lot·패널 위치별 격리, 보드 굴곡·분리·프로브 지지 "
                             "조건을 재현하고 공정 조건 수정",
    },
    "alternates_header": "조회 시각 2026-08-19 09:00:00   ·   데이터시트 "
                         "semiconductor_assembly_quality_datasheet_2026-08-18.xlsx "
                         "(기준일 2026-08-18)   ·   Group ID C-001 (MLCC)   ·   후보 2종",
    "relevance_header": "CRACK 관련 소견",
    "candidate_rows": [
        {"role": "주품목", "alternate_code": "—",
         "maker_pn": "Murata Manufacturing / GRM188R72A104KA35D",
         "key_spec": "MLCC / 0.10µF ±10%, 100V, X7R, 0603",
         "compatibility": "승인 주품목", "revalidation": "해당 없음",
         "relevance": "표준 단자 · 현재 CRACK 발생 중. 연성 단자(soft termination) 사양 아님",
         "source": "S-01 · 2026-08-18"},
        {"role": "대체 후보", "alternate_code": "ALT-01",
         "maker_pn": "KEMET / C0603C104K1RACTU",
         "key_spec": "MLCC / 0.10µF ±10%, 100V, X7R, 0603",
         "compatibility": "동급 후보 · 승인 전 사용 금지",
         "revalidation": "두께·단자·DC-bias·IR·land/reflow",
         "relevance": "표준 단자 — 굽힘 크랙 완충 없음. 대체해도 개선 근거 없음",
         "source": "S-02 · 2026-08-18"},
        {"role": "대체 후보", "alternate_code": "ALT-02",
         "maker_pn": "YAGEO / CC0603KRX7R0BB104",
         "key_spec": "MLCC / 0.10µF, 100V, X7R, 0603 후보",
         "compatibility": "동급 후보 · 승인 전 사용 금지",
         "revalidation": "두께·단자·DC-bias·IR·land/reflow",
         "relevance": "표준 단자 — 굽힘 크랙 완충 없음. 대체해도 개선 근거 없음",
         "source": "S-03 · 2026-08-18"},
    ],
}


# 종결본 샘플 — 담당자가 ①~⑥ 을 모두 회신하고 효과 검증까지 끝낸 상태.
# 상단 집계는 **발행 시점 스냅샷**이라 그대로 두고, 아래 회신과 검증만 채운다.
# 이 문서가 결재·감사에 남는 최종 형태다.
SAMPLE_CLOSED = dict(
    SAMPLE_DATA,
    alert_status="CLOSED",
    assignee="품질보증팀 김OO",
    support_teams="설비팀 · 생산팀 · 구매",
    footer="자동 발행   ·   샘플 (종결본 — ①~⑥ 회신 완료)   ·   결재 대상은 이 시트 1장",

    containment_scope="협조 필요    ☑ 생산 중단    ☑ 출하 보류    ☑ 재고 선별    "
                      "☐ 고객 통보    ☐ 설비 정지",
    reply_22=(
        "08-19 10:20  CAP-02·CAP-04 슬롯 정지, 생산팀장 승인(구두 10:15 · 결재 10:40).\n"
        "08-16 이후 생산분 전량 출하 보류 — job 7112~7139 / unit 10388~10604 (2,500대).\n"
        "전수 AOI 재검에서 9건 추가 검출, 누적 25건 / 24대 격리. 고객 통보 대상 아님(출하 전).\n"
        "배치 하강속도만 mock-v2 값으로 되돌린 임시 레시피 mock-v3h 적용 후 10:50 생산 재개. "
        "08-19 18:00 기준 재발 없음."
    ),
    reply_25="배치 하강속도 상향 + 지지 핀 최원거리 슬롯",
    reply_26=(
        "mock-v3 에서 배치 하강속도(18 → 32 mm/s)와 그리퍼 파지력(3.0 → 4.0 N)을 함께 올렸다.\n"
        "CAP-02·CAP-04 는 보드 지지 핀에서 가장 먼 위치라 착지 충격 시 국부 굽힘이 가장 크다. "
        "설비팀 로그에서 두 슬롯 피크력이 다른 CAP 슬롯의 1.7배(6.8N vs 4.0N).\n"
        "mock-v3h 로 되돌린 뒤 재발이 멈춘 것으로 검증했다(가설 아님). "
        "불량 25건이 입고 Lot 4개에 고르게 분포해 부품 Lot 요인은 배제한다."
    ),
    reply_28=(
        "인라인 AOI 판정 기준이 단자 들뜸과 외관 결손만 보고 있어 미세 굽힘 크랙을 걸러내지 못했다.\n"
        "「판단자료」 D 의 MLCC 항목은 '크랙 의심 Lot 단면 또는 비파괴 검사'를 요구하지만, "
        "현재 인라인에는 해당 항목이 없고 단면 검사는 주 1회 샘플로만 돌고 있었다.\n"
        "그 결과 08-16 생산분에 이미 크랙이 있었으나 08-18 16:30 임계 초과 시점까지 검출되지 않았다."
    ),
    reply_31=(
        "영구대책 ① 레시피 — 하강속도 20 mm/s, 파지력 3.2 N 으로 확정한 mock-v4 적용. "
        "변경관리 ECN-2608-114 승인 08-21, 적용 08-29.\n"
        "영구대책 ② 설비 — CAP-02·CAP-04 에 보드 지지 핀 2개 추가(설비팀, 08-28 완료). "
        "수평전개로 IND-01·IND-02 에도 함께 적용했다.\n"
        "부품 교체는 하지 않는다. 「대체품」 후보 2종 모두 표준 단자라 굽힘 크랙 개선 근거가 없다."
    ),
    verify_period="08-29 ~ 09-12",
    verify_recipe="mock-v4",
    verify_inspected=24500,
    verify_defective=3,
    verify_status="EFFECTIVE",
    verify_note=(
        "CRACK 3건 / 24,500건 = 122 PPM. 기준 400 PPM 의 0.31배이고 최소 검사 건수를 충족한다. "
        "대책 전 1,280 PPM 대비 10.5배 개선, CAP-02·CAP-04 집중도도 해소됐다."
    ),
    closed_at="2026-09-15",
    closure_note=(
        "영구대책 적용과 효과가 확인되어 종결한다. 잔여 위험 — 지지 핀은 현재 보드 리비전 기준이라 "
        "레이아웃 변경 시 재검토가 필요하다. AOI 기준 보완은 별건 QA-PROC-2609-003 으로 이관."
    ),
)


def build(data, path):
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    build_cover(wb, data)
    build_verify(wb, data)
    build_evidence(wb, data)
    build_analysis(wb, data)
    build_alternates(wb, data)

    for name, sheet, ref, _title, _prompt in REPLY_CELLS:
        wb.defined_names.add(
            DefinedName(name, attr_text=f"'{sheet}'!${ref[0]}${ref[1:]}"))

    wb.properties.title = "불량대책서"
    wb.properties.creator = "MAIN_SERVER/generate_defect_reports.py"
    wb.properties.created = dt.datetime(2026, 8, 19, 9, 0, 0)
    wb.save(path)
    print(f"wrote {path}")


if __name__ == "__main__":
    build(TOKEN_DATA, TEMPLATE_PATH)
    build(SAMPLE_DATA, SAMPLE_PATH)
    build(SAMPLE_CLOSED, CLOSED_PATH)
